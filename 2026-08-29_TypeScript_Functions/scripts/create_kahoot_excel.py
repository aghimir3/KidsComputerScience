"""Create the August 29 TypeScript functions Kahoot import workbook."""

import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


QUESTIONS = [
    (
        "What is the main purpose of a function?",
        "Store only one number",
        "Give reusable instructions a name",
        "Replace every TypeScript type",
        "Run a program without code",
        20,
        2,
    ),
    (
        "What does a function definition do?",
        "Describes the function's job",
        "Runs the function three times",
        "Deletes the function body",
        "Changes every argument to text",
        20,
        1,
    ),
    (
        "Which line is a function call?",
        "function cheer() {",
        'console.log("Go!");',
        "cheer();",
        "name: string",
        20,
        3,
    ),
    (
        "A function is defined but never called. What happens?",
        "Its body runs once",
        "Its body runs forever",
        "Its body does not run",
        "TypeScript calls it automatically",
        20,
        3,
    ),
    (
        "What is a parameter?",
        "A named input in a function definition",
        "The final line in every loop",
        "A file saved by TypeScript",
        "An error message from VS Code",
        20,
        1,
    ),
    (
        "What is an argument?",
        "The name of every function",
        "A value sent by a function call",
        "A value that must be an array",
        "The braces around a function body",
        20,
        2,
    ),
    (
        'In greet("Maya"), what is "Maya"?',
        "A function name",
        "A parameter type",
        "An argument",
        "A return type",
        20,
        3,
    ),
    (
        "What does name: string tell TypeScript?",
        "name must receive a string",
        "name must receive a number",
        "The function returns a string",
        "The function has no input",
        20,
        1,
    ),
    (
        "How many times does the body run after three calls?",
        "0",
        "1",
        "2",
        "3",
        20,
        4,
    ),
    (
        "Why does argument order matter?",
        "Calls are read from bottom to top",
        "Arguments match parameters by position",
        "TypeScript sorts arguments alphabetically",
        "The final argument becomes the function name",
        30,
        2,
    ),
    (
        "What is wrong with greetPlayer(42) for name: string?",
        "The function needs two names",
        "The argument has the wrong type",
        "The function name is too long",
        "Numbers cannot appear in programs",
        30,
        2,
    ),
    (
        "What can a function body contain?",
        "Only one console.log",
        "Only string values",
        "Familiar code such as decisions and loops",
        "No TypeScript instructions",
        20,
        3,
    ),
    (
        "What does return do?",
        "Sends one result back to the caller",
        "Prints every argument automatically",
        "Repeats the function forever",
        "Renames the function",
        30,
        1,
    ),
    (
        "Why use one review function for many AI labels?",
        "It guarantees every AI label is true",
        "It applies one consistent rule repeatedly",
        "It removes the need to inspect data",
        "It turns every label into a number",
        30,
        2,
    ),
    (
        "A review function flags unknown. What should happen next?",
        "A person inspects the original example",
        "The label is trusted automatically",
        "The entire program is deleted",
        "The function guesses a random label",
        20,
        1,
    ),
]


def main():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Kahoot Questions"
    headers = [
        "Question",
        "Answer 1",
        "Answer 2",
        "Answer 3",
        "Answer 4",
        "Time limit",
        "Correct answer",
    ]
    sheet.append(headers)
    for question in QUESTIONS:
        sheet.append(question)

    header_fill = PatternFill("solid", fgColor="1E3A5F")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    widths = [52, 34, 34, 34, 34, 14, 16]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = False
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        row[5].alignment = Alignment(horizontal="center", vertical="center")
        row[6].alignment = Alignment(horizontal="center", vertical="center")

    lesson_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    output_path = os.path.join(lesson_dir, "2026-08-29_Kahoot_Import.xlsx")
    workbook.save(output_path)
    print(f"Created {output_path}")
    print(f"Questions: {len(QUESTIONS)}")


if __name__ == "__main__":
    main()
