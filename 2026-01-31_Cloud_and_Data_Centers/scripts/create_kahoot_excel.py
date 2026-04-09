"""
Generate Kahoot Import Excel File
Topic: The Cloud & Data Centers
Date: January 31, 2026

Format: Question, Answer 1, Answer 2, Answer 3, Answer 4, Time limit, Correct answer(s)
Based on official Kahoot template format.

Run: python create_kahoot_excel.py
"""

import os

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    USE_OPENPYXL = True
except ImportError:
    USE_OPENPYXL = False

try:
    import xlsxwriter
    USE_XLSXWRITER = True
except ImportError:
    USE_XLSXWRITER = False


# Kahoot Questions Data
# Format: (Question, Answer1, Answer2, Answer3, Answer4, TimeLimit, CorrectAnswer)
# CorrectAnswer is 1-4 corresponding to Answer columns

QUESTIONS = [
    # === WHAT IS THE CLOUD? (Questions 1-5) ===
    (
        'What is "the cloud" in computing?',
        "Actual clouds in the sky",
        "Servers stored in data centers",
        "A type of Wi-Fi connection",
        "A wireless keyboard",
        30,
        2  # Answer 2 is correct
    ),
    (
        "There's a famous saying about the cloud. What is it?",
        "The cloud is the future",
        "Clouds are made of water",
        "There is no cloud, it's just someone else's computer",
        "The cloud never breaks",
        30,
        3  # Answer 3 is correct
    ),
    (
        "Where does your data actually go when you save it to the cloud?",
        "Into the atmosphere",
        "To a server in a data center",
        "It stays on your phone only",
        "Into outer space",
        30,
        2  # Answer 2 is correct
    ),
    (
        "What is NOT true about the cloud?",
        "It stores data on remote servers",
        "It allows access from any device",
        "It floats in the sky",
        "It requires internet to access",
        30,
        3  # Answer 3 is correct
    ),
    (
        "Which of these is an example of cloud storage?",
        "A USB flash drive",
        "Google Drive",
        "Your computer's hard drive",
        "A DVD",
        20,
        2  # Answer 2 is correct
    ),

    # === DATA CENTERS (Questions 6-10) ===
    (
        "What is a data center?",
        "A library for books",
        "A building with thousands of servers",
        "A type of computer monitor",
        "A place to charge phones",
        30,
        2  # Answer 2 is correct
    ),
    (
        "How many servers does Google have worldwide?",
        "About 100",
        "About 10,000",
        "Over 2 million",
        "Just 1 really big one",
        30,
        3  # Answer 3 is correct
    ),
    (
        "Why do data centers need massive cooling systems?",
        "To keep workers comfortable",
        "Servers generate a lot of heat",
        "To make ice for drinks",
        "They don't need cooling",
        30,
        2  # Answer 2 is correct
    ),
    (
        "Why are many data centers built near water?",
        "For drinking water",
        "For fishing",
        "Water helps with cooling",
        "They need to be near beaches",
        30,
        3  # Answer 3 is correct
    ),
    (
        "A single data center can use as much electricity as...",
        "A single house",
        "A small city",
        "A light bulb",
        "A bicycle",
        20,
        2  # Answer 2 is correct
    ),

    # === CLOUD SERVICES (Questions 11-14) ===
    (
        "Which of these is a cloud STORAGE service?",
        "Microsoft Word",
        "Calculator app",
        "iCloud",
        "Notepad",
        20,
        3  # Answer 3 is correct
    ),
    (
        "When you watch Netflix, where are the movies stored?",
        "On your TV",
        "On Netflix's servers in data centers",
        "On a DVD inside Netflix",
        "On your neighbor's computer",
        30,
        2  # Answer 2 is correct
    ),
    (
        "What happens to your Spotify music if you have no internet?",
        "It plays normally",
        "It gets louder",
        "You can only play downloaded songs",
        "Spotify deletes itself",
        30,
        3  # Answer 3 is correct
    ),
    (
        "When you save your Fortnite progress, where does it go?",
        "Your console only",
        "To cloud servers",
        "Into the game disk",
        "Nowhere - it's not saved",
        30,
        2  # Answer 2 is correct
    ),

    # === THE BIG THREE (Questions 15-17) ===
    (
        "Which company runs AWS (Amazon Web Services)?",
        "Google",
        "Microsoft",
        "Amazon",
        "Apple",
        20,
        3  # Answer 3 is correct
    ),
    (
        "Which is the BIGGEST cloud provider in the world?",
        "Google Cloud",
        "Microsoft Azure",
        "Amazon Web Services (AWS)",
        "Apple iCloud",
        30,
        3  # Answer 3 is correct
    ),
    (
        "Which company runs Azure cloud services?",
        "Amazon",
        "Microsoft",
        "Facebook",
        "Netflix",
        20,
        2  # Answer 2 is correct
    ),

    # === AI + CLOUD (Questions 18-24) ===
    (
        "Why does AI like ChatGPT need cloud computing?",
        "It doesn't need cloud",
        "AI needs massive computing power from many servers",
        "AI only works on phones",
        "Cloud makes AI slower",
        30,
        2  # Answer 2 is correct
    ),
    (
        "Which cloud provider runs ChatGPT?",
        "Google Cloud",
        "Amazon AWS",
        "Microsoft Azure",
        "Apple iCloud",
        20,
        3  # Answer 3 is correct
    ),
    (
        "Who created Grok AI?",
        "Google",
        "Microsoft",
        "xAI (Elon Musk)",
        "Amazon",
        20,
        3  # Answer 3 is correct
    ),
    (
        "Which AI chatbot is made by Anthropic?",
        "ChatGPT",
        "Grok",
        "Gemini",
        "Claude",
        20,
        4  # Answer 4 is correct
    ),
    (
        "What is Gemini?",
        "A video game",
        "Google's AI chatbot",
        "A social media app",
        "A type of server",
        20,
        2  # Answer 2 is correct
    ),
    (
        "Why can't powerful AI like GPT-4 run on just your phone?",
        "Phones don't have internet",
        "AI models are too big and need too much computing power",
        "AI doesn't work on mobile",
        "Phones are too expensive",
        30,
        2  # Answer 2 is correct
    ),
    (
        "What does AI need LOTS of to learn (train)?",
        "Water and food",
        "Data and computing power",
        "Sunlight",
        "Batteries",
        20,
        2  # Answer 2 is correct
    ),

    # === CLOUD SAFETY (Questions 25-26) ===
    (
        "What is the BEST way to protect your cloud accounts?",
        "Use 'password123'",
        "Never use the cloud",
        "Use strong passwords and 2FA",
        "Share your password with friends",
        30,
        3  # Answer 3 is correct
    ),
    (
        "What is an advantage of LOCAL storage over cloud storage?",
        "Need internet to access",
        "Works offline - no internet needed",
        "Can only use one device",
        "More expensive",
        30,
        2  # Answer 2 is correct
    ),
]


def create_with_openpyxl(output_path):
    """Create Excel file using openpyxl."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Kahoot Quiz"

    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    cell_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    correct_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers (exact Kahoot format)
    headers = [
        "Question",
        "Answer 1",
        "Answer 2",
        "Answer 3",
        "Answer 4",
        "Time limit (sec)",
        "Correct answer(s)"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Add questions
    for row_idx, q in enumerate(QUESTIONS, 2):
        question, a1, a2, a3, a4, time_limit, correct = q

        ws.cell(row=row_idx, column=1, value=question).alignment = cell_alignment
        ws.cell(row=row_idx, column=2, value=a1).alignment = cell_alignment
        ws.cell(row=row_idx, column=3, value=a2).alignment = cell_alignment
        ws.cell(row=row_idx, column=4, value=a3).alignment = cell_alignment
        ws.cell(row=row_idx, column=5, value=a4).alignment = cell_alignment
        ws.cell(row=row_idx, column=6, value=time_limit).alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=7, value=correct).alignment = Alignment(horizontal="center")

        # Add borders to all cells
        for col in range(1, 8):
            ws.cell(row=row_idx, column=col).border = thin_border

        # Highlight correct answer column
        correct_col = correct + 1  # +1 because answers start at column 2
        ws.cell(row=row_idx, column=correct_col).fill = correct_fill

    # Set column widths
    ws.column_dimensions['A'].width = 55  # Question
    ws.column_dimensions['B'].width = 35  # Answer 1
    ws.column_dimensions['C'].width = 35  # Answer 2
    ws.column_dimensions['D'].width = 35  # Answer 3
    ws.column_dimensions['E'].width = 35  # Answer 4
    ws.column_dimensions['F'].width = 15  # Time limit
    ws.column_dimensions['G'].width = 15  # Correct answer

    # Set row height for header
    ws.row_dimensions[1].height = 25

    wb.save(output_path)
    return True


def create_with_xlsxwriter(output_path):
    """Create Excel file using xlsxwriter."""
    workbook = xlsxwriter.Workbook(output_path)
    worksheet = workbook.add_worksheet("Kahoot Quiz")

    # Define formats
    header_format = workbook.add_format({
        'bold': True,
        'font_color': 'white',
        'bg_color': '#1E3A5F',
        'align': 'center',
        'valign': 'vcenter',
        'text_wrap': True,
        'border': 1
    })

    cell_format = workbook.add_format({
        'align': 'left',
        'valign': 'vcenter',
        'text_wrap': True,
        'border': 1
    })

    center_format = workbook.add_format({
        'align': 'center',
        'valign': 'vcenter',
        'border': 1
    })

    correct_format = workbook.add_format({
        'align': 'left',
        'valign': 'vcenter',
        'text_wrap': True,
        'border': 1,
        'bg_color': '#C6EFCE'
    })

    # Headers
    headers = [
        "Question",
        "Answer 1",
        "Answer 2",
        "Answer 3",
        "Answer 4",
        "Time limit (sec)",
        "Correct answer(s)"
    ]

    for col, header in enumerate(headers):
        worksheet.write(0, col, header, header_format)

    # Add questions
    for row_idx, q in enumerate(QUESTIONS, 1):
        question, a1, a2, a3, a4, time_limit, correct = q

        worksheet.write(row_idx, 0, question, cell_format)

        # Write answers, highlighting correct one
        answers = [a1, a2, a3, a4]
        for i, answer in enumerate(answers):
            fmt = correct_format if (i + 1) == correct else cell_format
            worksheet.write(row_idx, i + 1, answer, fmt)

        worksheet.write(row_idx, 5, time_limit, center_format)
        worksheet.write(row_idx, 6, correct, center_format)

    # Set column widths
    worksheet.set_column(0, 0, 55)   # Question
    worksheet.set_column(1, 4, 35)   # Answers
    worksheet.set_column(5, 6, 15)   # Time & Correct

    # Set row height for header
    worksheet.set_row(0, 25)

    workbook.close()
    return True


def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    parent_dir = os.path.dirname(script_dir)
    output_path = os.path.join(parent_dir, "2026-01-31_Kahoot_Import.xlsx")

    print("Creating Kahoot import Excel file...")
    print(f"  Topic: The Cloud & Data Centers")
    print(f"  Questions: {len(QUESTIONS)}")

    success = False

    if USE_OPENPYXL:
        print("  Using: openpyxl")
        success = create_with_openpyxl(output_path)
    elif USE_XLSXWRITER:
        print("  Using: xlsxwriter")
        success = create_with_xlsxwriter(output_path)
    else:
        print("ERROR: Neither openpyxl nor xlsxwriter is installed.")
        print("Install one with: pip install openpyxl")
        return

    if success:
        print(f"\n[SUCCESS] Kahoot Excel file created: {output_path}")
        print("\nTo import into Kahoot:")
        print("  1. Go to kahoot.com and click 'Create'")
        print("  2. Click 'Add question' > 'Import'")
        print("  3. Select 'Import spreadsheet'")
        print("  4. Upload this .xlsx file")


if __name__ == "__main__":
    main()
