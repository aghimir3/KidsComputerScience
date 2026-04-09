"""
Generate Kahoot Import Excel File
Format: Question, Answer 1, Answer 2, Answer 3, Answer 4, Time limit, Correct answer(s)

Based on official Kahoot template format.
Run: python create_kahoot_excel.py
"""

import os

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    USE_OPENPYXL = True
except ImportError:
    USE_OPENPYXL = False

try:
    import xlsxwriter
    USE_XLSXWRITER = True
except ImportError:
    USE_XLSXWRITER = False


# Kahoot Questions Data
# Format: (Question, Answer1, Answer2, Answer3, Answer4, TimeLimit, CorrectAnswer)
# CorrectAnswer is 1-4 corresponding to Answer columns

QUESTIONS = [
    # === CLIENTS & SERVERS (Questions 1-5) ===
    (
        "What is a CLIENT in networking?",
        "A computer that stores websites",
        "A device that ASKS for information",
        "The internet's phone book",
        "A type of cable",
        30,
        2  # Answer 2 is correct
    ),
    (
        "What is a SERVER?",
        "Your personal laptop",
        "A device that asks for websites",
        "A computer that SERVES information to clients",
        "A type of Wi-Fi connection",
        30,
        3  # Answer 3 is correct
    ),
    (
        "When you visit google.com, YOUR device is the...",
        "Server",
        "Router",
        "Client",
        "Gateway",
        20,
        3  # Answer 3 is correct
    ),
    (
        "Which of these is an example of a SERVER?",
        "Your smartphone",
        "Your gaming console",
        "A computer at Google storing YouTube videos",
        "Your laptop at home",
        30,
        3  # Answer 3 is correct
    ),
    (
        "When you send a message to a server, that's called a...",
        "Response",
        "Download",
        "Request",
        "Upload",
        20,
        3  # Answer 3 is correct
    ),

    # === HTTP & HTTPS (Questions 6-10) ===
    (
        "What does HTTP stand for?",
        "Hyper Text Transfer Protocol",
        "High Tech Transfer Program",
        "Home Text Transfer Protocol",
        "Hyper Transfer Text Program",
        30,
        1  # Answer 1 is correct
    ),
    (
        'What does the "S" in HTTPS stand for?',
        "Speed",
        "Server",
        "Secure",
        "Simple",
        20,
        3  # Answer 3 is correct
    ),
    (
        "HTTPS is like sending a message in a...",
        "Postcard (anyone can read it)",
        "Locked box (only you and the server can read it)",
        "Newspaper (everyone sees it)",
        "Megaphone (everyone hears it)",
        30,
        2  # Answer 2 is correct
    ),
    (
        "When should you ALWAYS check for HTTPS?",
        "When watching videos",
        "When reading news articles",
        "When entering passwords or credit card info",
        "When looking at pictures",
        30,
        3  # Answer 3 is correct
    ),
    (
        "What does the lock icon in your browser mean?",
        "The website is loading slowly",
        "The connection is encrypted (HTTPS)",
        "The website has no ads",
        "Your computer is locked",
        20,
        2  # Answer 2 is correct
    ),

    # === PORTS (Questions 11-14) ===
    (
        "What is a PORT in networking?",
        "A type of cable",
        "Like an apartment number for services on a computer",
        "A wireless connection",
        "The same as an IP address",
        30,
        2  # Answer 2 is correct
    ),
    (
        "What port does HTTPS (secure web) use?",
        "Port 80",
        "Port 25",
        "Port 443",
        "Port 22",
        20,
        3  # Answer 3 is correct
    ),
    (
        "What port does regular HTTP use?",
        "Port 443",
        "Port 80",
        "Port 22",
        "Port 53",
        20,
        2  # Answer 2 is correct
    ),
    (
        "If the IP address is like a building address, the PORT is like...",
        "The street name",
        "The city name",
        "The apartment number",
        "The zip code",
        20,
        3  # Answer 3 is correct
    ),

    # === DNS & URLs (Questions 15-18) ===
    (
        "What does DNS stand for?",
        "Digital Network Service",
        "Domain Name System",
        "Data Number Storage",
        "Direct Network Server",
        30,
        2  # Answer 2 is correct
    ),
    (
        "What does DNS do?",
        "Blocks dangerous websites",
        "Speeds up your internet",
        "Translates domain names to IP addresses",
        "Stores your passwords",
        30,
        3  # Answer 3 is correct
    ),
    (
        'In the URL "https://www.google.com/search", what is "google"?',
        "The protocol",
        "The domain name",
        "The path",
        "The port",
        30,
        2  # Answer 2 is correct
    ),
    (
        'In the URL "https://www.google.com/search", what is "/search"?',
        "The protocol",
        "The domain name",
        "The path",
        "The port",
        30,
        3  # Answer 3 is correct
    ),

    # === PUTTING IT ALL TOGETHER (Questions 19-20) ===
    (
        "What happens FIRST when you type google.com in your browser?",
        "Google's server sends you the webpage",
        "Your browser shows the page",
        "Your computer asks DNS for the IP address",
        "Packets travel through the internet",
        30,
        3  # Answer 3 is correct
    ),
    (
        'What does the "tracert" command show you?',
        "Your browsing history",
        "Every stop (hop) your data makes on its journey",
        "Your computer's password",
        "How fast your internet is",
        30,
        2  # Answer 2 is correct
    ),
]


def create_with_openpyxl(output_path):
    """Create Excel file using openpyxl."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Kahoot Quiz"

    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    cell_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    correct_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers (exact Kahoot format)
    headers = [
        "Question",
        "Answer 1",
        "Answer 2",
        "Answer 3",
        "Answer 4",
        "Time limit (sec)",
        "Correct answer(s)"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Add questions
    for row_idx, q in enumerate(QUESTIONS, 2):
        question, a1, a2, a3, a4, time_limit, correct = q

        ws.cell(row=row_idx, column=1, value=question).alignment = cell_alignment
        ws.cell(row=row_idx, column=2, value=a1).alignment = cell_alignment
        ws.cell(row=row_idx, column=3, value=a2).alignment = cell_alignment
        ws.cell(row=row_idx, column=4, value=a3).alignment = cell_alignment
        ws.cell(row=row_idx, column=5, value=a4).alignment = cell_alignment
        ws.cell(row=row_idx, column=6, value=time_limit).alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=7, value=correct).alignment = Alignment(horizontal="center")

        # Add borders to all cells
        for col in range(1, 8):
            ws.cell(row=row_idx, column=col).border = thin_border

        # Highlight correct answer column
        correct_col = correct + 1  # +1 because answers start at column 2
        ws.cell(row=row_idx, column=correct_col).fill = correct_fill

    # Set column widths
    ws.column_dimensions['A'].width = 55  # Question
    ws.column_dimensions['B'].width = 35  # Answer 1
    ws.column_dimensions['C'].width = 35  # Answer 2
    ws.column_dimensions['D'].width = 35  # Answer 3
    ws.column_dimensions['E'].width = 35  # Answer 4
    ws.column_dimensions['F'].width = 15  # Time limit
    ws.column_dimensions['G'].width = 15  # Correct answer

    # Set row height for header
    ws.row_dimensions[1].height = 25

    wb.save(output_path)
    return True


def create_with_xlsxwriter(output_path):
    """Create Excel file using xlsxwriter."""
    workbook = xlsxwriter.Workbook(output_path)
    worksheet = workbook.add_worksheet("Kahoot Quiz")

    # Define formats
    header_format = workbook.add_format({
        'bold': True,
        'font_color': 'white',
        'bg_color': '#1E3A5F',
        'align': 'center',
        'valign': 'vcenter',
        'text_wrap': True,
        'border': 1
    })

    cell_format = workbook.add_format({
        'align': 'left',
        'valign': 'vcenter',
        'text_wrap': True,
        'border': 1
    })

    center_format = workbook.add_format({
        'align': 'center',
        'valign': 'vcenter',
        'border': 1
    })

    correct_format = workbook.add_format({
        'align': 'left',
        'valign': 'vcenter',
        'text_wrap': True,
        'border': 1,
        'bg_color': '#C6EFCE'
    })

    # Headers
    headers = [
        "Question",
        "Answer 1",
        "Answer 2",
        "Answer 3",
        "Answer 4",
        "Time limit (sec)",
        "Correct answer(s)"
    ]

    for col, header in enumerate(headers):
        worksheet.write(0, col, header, header_format)

    # Add questions
    for row_idx, q in enumerate(QUESTIONS, 1):
        question, a1, a2, a3, a4, time_limit, correct = q

        worksheet.write(row_idx, 0, question, cell_format)

        # Write answers, highlighting correct one
        answers = [a1, a2, a3, a4]
        for i, answer in enumerate(answers):
            fmt = correct_format if (i + 1) == correct else cell_format
            worksheet.write(row_idx, i + 1, answer, fmt)

        worksheet.write(row_idx, 5, time_limit, center_format)
        worksheet.write(row_idx, 6, correct, center_format)

    # Set column widths
    worksheet.set_column(0, 0, 55)   # Question
    worksheet.set_column(1, 4, 35)   # Answers
    worksheet.set_column(5, 6, 15)   # Time & Correct

    # Set row height for header
    worksheet.set_row(0, 25)

    workbook.close()
    return True


def main():
    output_path = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        "2026-01-25_Kahoot_Import.xlsx"
    )

    print("Creating Kahoot import Excel file...")
    print(f"  Questions: {len(QUESTIONS)}")

    success = False

    if USE_OPENPYXL:
        print("  Using: openpyxl")
        success = create_with_openpyxl(output_path)
    elif USE_XLSXWRITER:
        print("  Using: xlsxwriter")
        success = create_with_xlsxwriter(output_path)
    else:
        print("ERROR: Neither openpyxl nor xlsxwriter is installed.")
        print("Install one with: pip install openpyxl")
        return

    if success:
        print(f"\n[SUCCESS] Kahoot Excel file created: {output_path}")
        print("\nTo import into Kahoot:")
        print("  1. Go to kahoot.com and click 'Create'")
        print("  2. Click 'Add question' > 'Import'")
        print("  3. Select 'Import spreadsheet'")
        print("  4. Upload this .xlsx file")


if __name__ == "__main__":
    main()
