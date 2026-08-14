"""Create the August 15 arrays and loops Kahoot import workbook."""

import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


QUESTIONS = [
    (
        "What does an array store?",
        "One ordered list of values",
        "Only the final loop result",
        "One TypeScript file name",
        "Only true or false",
        20,
        1,
    ),
    (
        "What does string[] mean?",
        "One string with square brackets",
        "An array whose items are strings",
        "An array whose items are numbers",
        "A loop that prints strings",
        20,
        2,
    ),
    (
        "What is the first valid index in an array?",
        "-1",
        "0",
        "1",
        "It depends on the array length",
        20,
        2,
    ),
    (
        "Which index selects the second item?",
        "0",
        "1",
        "2",
        "3",
        20,
        2,
    ),
    (
        'What does colors[2] select from ["red", "green", "blue"]?',
        "red",
        "green",
        "blue",
        "2",
        20,
        3,
    ),
    (
        "What is the length of [10, 20, 30, 40]?",
        "3",
        "4",
        "10",
        "40",
        20,
        2,
    ),
    (
        "An array has a length of 5. What is its last valid index?",
        "3",
        "4",
        "5",
        "6",
        20,
        2,
    ),
    (
        "Which expression selects the final item in players?",
        "players[players.length]",
        "players[players.length + 1]",
        "players[players.length - 1]",
        "players[0 - players.length]",
        30,
        3,
    ),
    (
        "Which loop condition visits every valid index once?",
        "i <= items.length",
        "i < items.length",
        "i > items.length",
        "i === items.length",
        20,
        2,
    ),
    (
        "Why can i <= items.length produce undefined?",
        "It stops before the first item",
        "It changes every item into text",
        "It tries the index equal to the length",
        "It makes the array empty",
        30,
        3,
    ),
    (
        "What are the two jobs of i in an array loop?",
        "Store the array and print its length",
        "Count rounds and select the current item",
        "Change strings into numbers",
        "Start and stop VS Code",
        30,
        2,
    ),
    (
        "What does scores[1] = 88 change?",
        "The first score",
        "The second score",
        "Every score",
        "The array length",
        20,
        2,
    ),
    (
        "Where should an accumulator be created?",
        "Before the loop",
        "Inside the loop after every print",
        "Inside the array brackets",
        "After the final output",
        20,
        1,
    ),
    (
        "Why might an AI project process an array of labels?",
        "To store and inspect a collection of examples",
        "To guarantee every label is correct",
        "To remove the need for human decisions",
        "To turn VS Code into an AI model",
        30,
        1,
    ),
    (
        "A checker finds a label marked unknown. What should happen next?",
        "Trust it as correct automatically",
        "Delete the complete dataset",
        "Ask a human to review the example",
        "Stop using arrays",
        20,
        3,
    ),
]


def main():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Kahoot Questions"
    headers = [
        "Question",
        "Answer 1",
        "Answer 2",
        "Answer 3",
        "Answer 4",
        "Time limit",
        "Correct answer",
    ]
    sheet.append(headers)
    for question in QUESTIONS:
        sheet.append(question)

    header_fill = PatternFill("solid", fgColor="1E3A5F")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    widths = [52, 34, 34, 34, 34, 14, 16]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = False
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        row[5].alignment = Alignment(horizontal="center", vertical="center")
        row[6].alignment = Alignment(horizontal="center", vertical="center")

    lesson_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    output_path = os.path.join(lesson_dir, "2026-08-15_Kahoot_Import.xlsx")
    workbook.save(output_path)
    print(f"Created {output_path}")
    print(f"Questions: {len(QUESTIONS)}")


if __name__ == "__main__":
    main()